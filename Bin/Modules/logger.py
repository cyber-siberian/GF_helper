import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import datetime
import shutil
from win32com.client.gencache import EnsureDispatch
from win32com.client import constants
import os

if not os.path.exists('Bin/Modules/utils'):
    os.makedirs('Bin/Modules/utils')
flowloger = os.getcwd() + '/Bin/Modules/utils/flowloger.xlsx'

class Logger():
    def __init__(self, html=None):
        self.HTML_DOC_PATH = html
        try:
            self.xwb = openpyxl.load_workbook(flowloger)
        except FileNotFoundError:
            self.xwb = openpyxl.Workbook()
            self.xwb.save(flowloger)
        self.xsheet = self.xwb['Sheet']

    def save_document(self):
        try:
            self.xwb.save(flowloger)
            self.update_html()
            print('[+]Log file is updated')
            return True
        except PermissionError as ex:
            print("[!]Can't save .xlsx file -> " + str(ex))
            print("[!]Can't save '" + flowloger + "' file -> " + str(ex))
            return False

    def set_current_date(self):
        last_row = self.xsheet.max_row
        date_cell = self.xsheet.cell(row=last_row+2, column=1)
        date = (str(datetime.date.today().day) if datetime.date.today().day > 9 else '0' + str(datetime.date.today().day)) + "." \
               + (str(datetime.date.today().month) if datetime.date.today().month > 9 else '0' + str(datetime.date.today().month))  + "." \
               + str(datetime.date.today().year)
        for i in range(1, last_row+1):
            if self.xsheet.cell(row=i, column=1).value == date:
                return True
        date_cell.value = date
        dateFill = PatternFill(start_color='e0edff',
                               end_color='e0edff',
                               fill_type='solid')
        dateBorder = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        date_cell.fill = dateFill
        date_cell.border = dateBorder
        val = date_cell.value
        self.save_document()
        print("[+]Current date set to " + val)
        return True

    def set_new_task(self, task, price):
        self.set_current_date()
        last_row = 1
        last_date = 1
        for i in reversed(range(1, self.xsheet.max_row+1)):
            if self.xsheet.cell(row=i, column=1).value is not None:
                last_date = i
                break
        for i in range(last_date, self.xsheet.max_row+2):
            if self.xsheet.cell(row=i, column=2).value is None:
                last_row = i
                break
        for i in range(last_date, last_row):
            if self.xsheet.cell(row=i, column=2).value == str(task):
                print("[!]Task " + task + " is already in list -> [B" + str(i) + "]")
                return False
        task_cell = self.xsheet.cell(row=last_row, column=2)
        price_cell = self.xsheet.cell(row=last_row, column=3)
        cFill = PatternFill(start_color='d77bdd',
                               end_color='d77bdd',
                               fill_type='solid')
        cBorder = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        task_cell.fill = price_cell.fill = cFill
        task_cell.border = price_cell.border = cBorder
        task_cell.value = str(task)
        price_cell.value = str(price)
        self.save_document()
        print("[+]New task " + task + " (" + price + ") added")
        return True

    def set_task_done(self, task):
        date = (str(datetime.date.today().day) if datetime.date.today().day > 9
                else '0' + str(datetime.date.today().day)) + "." \
               + (str(datetime.date.today().month) if datetime.date.today().month > 9
                  else '0' + str(datetime.date.today().month)) + "." \
               + str(datetime.date.today().year)
        last_row = self.xsheet.max_row
        for i in range(1, last_row+1):
            if self.xsheet.cell(row=i, column=2).value == str(task):
                cFill = PatternFill(start_color='4ef050',
                                    end_color='4ef050',
                                    fill_type='solid')
                cBorder = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
                self.xsheet.cell(row=i, column=4).value = str(date)
                self.xsheet.cell(row=i, column=2).fill = cFill
                self.xsheet.cell(row=i, column=3).fill = cFill
                self.xsheet.cell(row=i, column=4).fill = cFill
                self.xsheet.cell(row=i, column=2).border = cBorder
                self.xsheet.cell(row=i, column=3).border = cBorder
                self.xsheet.cell(row=i, column=4).border = cBorder
                self.save_document()
                print("[+]Task " + str(task) + " set done on " + date)
                return True
        return False

    def set_task_missed(self, task):
        date = (str(datetime.date.today().day) if datetime.date.today().day > 9
                else '0' + str(datetime.date.today().day)) + "." \
               + (str(datetime.date.today().month) if datetime.date.today().month > 9
                  else '0' + str(datetime.date.today().month)) + "." \
               + str(datetime.date.today().year)
        last_row = self.xsheet.max_row
        for i in range(1, last_row+1):
            if self.xsheet.cell(row=i, column=2).value == str(task):
                cFill = PatternFill(start_color='ffff66',
                                    end_color='ffff66',
                                    fill_type='solid')
                cBorder = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
                self.xsheet.cell(row=i, column=4).value = str(date)
                self.xsheet.cell(row=i, column=2).fill = cFill
                self.xsheet.cell(row=i, column=3).fill = cFill
                self.xsheet.cell(row=i, column=4).fill = cFill
                self.xsheet.cell(row=i, column=2).border = cBorder
                self.xsheet.cell(row=i, column=3).border = cBorder
                self.xsheet.cell(row=i, column=4).border = cBorder
                self.save_document()
                print("[+]Task " + str(task) + " set missed on " + date)
                return True
        return False

    def update_html(self):
        xl = EnsureDispatch('Excel.Application')
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(flowloger)
        wb.SaveAs(self.HTML_DOC_PATH, constants.xlHtml)
        xl.Workbooks.Close()
        xl.Quit()
